import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from fuzzywuzzy import process
from datetime import date

st.title("Weekly Payroll Calculator with Closers & Enrollers")

# ---------- Helpers ----------
def parse_and_round_up(duration_str: str) -> float:
    """Parses time strings like 'H:M:S' and rounds the total hours up."""
    try:
        h, m, s = map(int, str(duration_str).split(':'))
        total_hours = h + m/60 + s/3600
        return float(np.ceil(total_hours))
    except (ValueError, TypeError):
        return 0.0

def fuzzy_match(name: str, name_list, cutoff=80) -> str:
    """Finds the best match for a name in a list, otherwise returns the original name."""
    match = process.extractOne(name, name_list, score_cutoff=cutoff)
    return match[0] if match else name

def determine_hourly_rate(deals: int) -> int:
    """Determines the closer's hourly rate based on their deal count."""
    if deals >= 15:
        return 22
    elif deals >= 12:
        return 20
    elif deals >= 8:
        return 18
    elif deals >= 4:
        return 15
    else:
        return 13

def hours_bonus(hours: float) -> int:
    """Calculates a bonus based on the total manual hours worked."""
    if hours >= 60:
        return 100
    elif hours >= 50:
        return 75
    elif hours >= 40:
        return 50
    else:
        return 0

# ---------- Inputs ----------
hubspot_file = st.file_uploader("Upload HubSpot Deal Tracker CSV")
closer_hours_file = st.file_uploader("Upload Closer Timesheet CSV")
enroller_hours_file = st.file_uploader("Upload Enroller Timesheet CSV")

if not (hubspot_file and closer_hours_file and enroller_hours_file):
    st.warning("Please upload all three CSV files to proceed.")
    st.stop()

# ---------- Load & Clean Data ----------
# HubSpot Deals
hubspot_df = pd.read_csv(hubspot_file)
hubspot_df.columns = hubspot_df.columns.str.strip()
hubspot_df['DATE'] = pd.to_datetime(hubspot_df['DATE'], errors='coerce')
hubspot_df['CLOSER'] = hubspot_df['CLOSER'].astype(str).str.strip().str.title()
hubspot_df['ENROLLER'] = hubspot_df['ENROLLER'].astype(str).str.strip().str.title()
# Create a dedicated column for date-only comparisons
hubspot_df['DealDate'] = hubspot_df['DATE'].dt.date

# Closer Hours
closer_df = pd.read_csv(closer_hours_file)
closer_df.columns = closer_df.columns.str.strip()
closer_df['Rep'] = closer_df['Rep'].astype(str).str.strip().str.title()
closer_df['Man Hours'] = closer_df['Man Hours'].astype(str).apply(parse_and_round_up)

# Enroller Hours
enroller_df = pd.read_csv(enroller_hours_file)
enroller_df.columns = enroller_df.columns.str.strip()
enroller_df['Rep'] = enroller_df['Rep'].astype(str).str.strip().str.title()
enroller_df['Man Hours'] = enroller_df['Man Hours'].astype(str).apply(parse_and_round_up)


# ---------- Closer Calculations ----------
st.header("Closer Payroll")

# Standardize Closer names against a single source of truth (HubSpot list)
canonical_closers = hubspot_df['CLOSER'].unique()
closer_df['Agent'] = closer_df['Rep'].apply(lambda x: fuzzy_match(x, canonical_closers))
hubspot_df['Agent'] = hubspot_df['CLOSER'].apply(lambda x: fuzzy_match(x, canonical_closers))

# Total deal counts
deal_counts = hubspot_df['Agent'].value_counts().rename_axis('Agent').reset_index(name='Deal Count')

# Automatically count deals on ANY Saturday in the file
saturday_deals_df = hubspot_df[hubspot_df['DATE'].dt.weekday == 5] # Monday=0, Saturday=5
saturday_deals = saturday_deals_df['Agent'].value_counts().rename_axis('Agent').reset_index(name='Saturday Deals')

# First deal of each day (company-wide) bonus
# **FIXED LINE:** Use the column name 'DealDate' in the subset list
first_per_date = hubspot_df.sort_values(by='DATE').drop_duplicates(subset=['DealDate'], keep='first')
first_deal_bonus = first_per_date['Agent'].value_counts().rename_axis('Agent').reset_index(name='First Deal Bonus Count')

# **NEW LOGIC: Start with agents from the timesheet file ONLY.**
# This ensures that no agent can have deals without having hours.
closers = closer_df[['Agent', 'Man Hours']].copy().drop_duplicates(subset=['Agent'])

# Merge all closer data together using a 'left' join to keep only the agents with hours
closers = closers.merge(deal_counts, on='Agent', how='left')
closers = closers.merge(saturday_deals, on='Agent', how='left')
closers = closers.merge(first_deal_bonus, on='Agent', how='left')
closers = closers.fillna(0) # Fill NaNs with 0 for agents who had hours but zero deals

# Compute pay components
closers['Hourly Rate'] = closers['Deal Count'].apply(determine_hourly_rate)
closers['Hourly Pay'] = closers['Hourly Rate'] * closers['Man Hours']
closers['Regular Deals'] = closers['Deal Count'] - closers['Saturday Deals']
closers['Regular Deals Pay'] = closers['Regular Deals'] * 35
closers['Saturday Deals Pay'] = closers['Saturday Deals'] * 50
closers['Hours Bonus'] = closers['Man Hours'].apply(hours_bonus)
closers['First Deal Bonus'] = closers['First Deal Bonus Count'] * 25


# ---------- Enroller Calculations ----------
st.header("Enroller Payroll")

# Standardize Enroller names
canonical_enrollers = hubspot_df['ENROLLER'].unique()
enroller_df['Agent'] = enroller_df['Rep'].apply(lambda x: fuzzy_match(x, canonical_enrollers))
hubspot_df['Enroller Agent'] = hubspot_df['ENROLLER'].apply(lambda x: fuzzy_match(x, canonical_enrollers))

# Count submissions
enroller_submissions = hubspot_df['Enroller Agent'].value_counts().rename_axis('Agent').reset_index(name='Submitted Deals')

# **NEW LOGIC: Apply the same rule to enrollers.**
# Start with agents from the timesheet file ONLY.
enrollers = enroller_df[['Agent', 'Man Hours']].copy().drop_duplicates(subset=['Agent'])

# Merge all enroller data
enrollers = enrollers.merge(enroller_submissions, on='Agent', how='left')
enrollers = enrollers.fillna(0)

# Enroller pay rules
enrollers['Hourly Rate'] = 18
enrollers['Hourly Pay'] = enrollers['Man Hours'] * enrollers['Hourly Rate']
enrollers['Regular Deals Pay'] = enrollers['Submitted Deals'] * 5
enrollers['Saturday Deals Pay'] = 0
enrollers['Hours Bonus'] = 0
enrollers['First Deal Bonus'] = 0
enrollers['Deal Count'] = enrollers['Submitted Deals']


# ---------- Prepare for Export ----------
# Select and order columns for closers
closers_export = closers[[
    'Agent', 'Deal Count', 'Man Hours', 'Hourly Rate', 'Hourly Pay',
    'Regular Deals Pay', 'Saturday Deals Pay', 'Hours Bonus', 'First Deal Bonus'
]].copy()

# Select and order columns for enrollers
enrollers_export = enrollers[[
    'Agent', 'Deal Count', 'Man Hours', 'Hourly Rate', 'Hourly Pay',
    'Regular Deals Pay', 'Saturday Deals Pay', 'Hours Bonus', 'First Deal Bonus'
]].copy()

# Add blank columns for manual entry in the XLSX file
for df in [closers_export, enrollers_export]:
    df['Manual Bonus'] = 0
    df['$25 Bonus Count'] = 0
    df['$50 Bonus Count'] = 0

# Sort each section alphabetically by Agent name
closers_export = closers_export.sort_values(by='Agent').reset_index(drop=True)
enrollers_export = enrollers_export.sort_values(by='Agent').reset_index(drop=True)

# Combine into a single DataFrame for the final report
combined_export = pd.concat([closers_export, enrollers_export], ignore_index=True).fillna(0)

# ---------- Create XLSX File with Formulas ----------
output = BytesIO()
wb = Workbook()
ws = wb.active
ws.title = "Payroll Summary"

# Define headers for the Excel file
headers = [
    'Agent', 'Deal Count', 'Man Hours', 'Hourly Rate', 'Hourly Pay',
    'Regular Deals Pay', 'Saturday Deals Pay', 'Hours Bonus', 'First Deal Bonus',
    'Manual Bonus', '$25 Bonus Count', '$50 Bonus Count', 'Total Pay', 'CPA'
]
ws.append(headers)

# Write data and formulas to the Excel sheet row by row
for idx, row in combined_export.iterrows():
    row_num = idx + 2  # Excel rows are 1-based, and we have a header
    ws.append([
        row['Agent'],
        row['Deal Count'],
        row['Man Hours'],
        row['Hourly Rate'],
        row['Hourly Pay'],
        row['Regular Deals Pay'],
        row['Saturday Deals Pay'],
        row['Hours Bonus'],
        row['First Deal Bonus'],
        '',  # Manual Bonus (editable)
        '',  # $25 Bonus Count (editable)
        '',  # $50 Bonus Count (editable)
        f"=E{row_num}+F{row_num}+G{row_num}+H{row_num}+I{row_num}+J{row_num}+(K{row_num}*25)+(L{row_num}*50)", # Total Pay Formula
        f"=IF(B{row_num}>0, M{row_num}/B{row_num}, 0)"  # CPA Formula with protection for zero deals
    ])

# Add Overall CPA calculation at the bottom
total_rows = len(combined_export) + 2
ws[f"L{total_rows}"] = "Overall CPA:"
ws[f"M{total_rows}"] = f"=SUM(M2:M{total_rows-1})/SUM(B2:B{total_rows-1})"

# Save workbook to a byte stream
wb.save(output)
output.seek(0) # Rewind the stream to the beginning

# ---------- Display Results and Download Link ----------
st.subheader("Preview of Payroll Data")
st.dataframe(combined_export.round(2))
st.caption("Agents without logged hours in the timesheet files will not be included in the payroll report.")


st.download_button(
    label="Download Payroll XLSX",
    data=output,
    file_name="Payroll_Summary.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
